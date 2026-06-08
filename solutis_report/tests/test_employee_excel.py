"""Tests for EmployeeExcelGenerator"""

import openpyxl
from services.employee_excel import EmployeeExcelGenerator


def test_get_headers():
    """Test that headers match the expected columns."""
    generator = EmployeeExcelGenerator()
    headers = generator.get_headers()

    assert headers == [
        "N°",
        "COLABORADOR",
        "CPF",
        "CARGO",
        "PROJETO",
        "BU",
        "CENTRO DE CUSTO",
        "CENTRO DE CUSTO (código)",
        "GESTOR",
        "EXECUTIVO",
        "LOCAL DE TRABALHO",
        "DESCRIÇÃO DO EQUIPAMENTO",
        "PATRIMÔNIO",
        "PADRÃO EQUIPAMENTO",
        "STATUS",
    ]


def test_get_sheet_title():
    """Test the sheet title."""
    generator = EmployeeExcelGenerator()
    assert generator.get_sheet_title() == "CONSULTA POR COLABORADOR"


def test_generate_excel_with_data():
    """Test Excel generation with sample data."""
    generator = EmployeeExcelGenerator()
    data = [
        {
            "employee": "João Silva",
            "code": "12345678901",
            "role": "Desenvolvedor",
            "project": "Projeto X",
            "bu": "TI",
            "cost_center": "Centro Custo 1",
            "cost_center_code": "CC001",
            "manager": "Gestor A",
            "executive": "Exec A",
            "workload": "Home Office",
            "equipment_description": "Notebook Dell",
            "patrimony": "PAT-001",
            "equipment_standard": "Notebook Padrão",
            "status": "Ativo",
        }
    ]

    excel_bytes = generator.generate(data)
    assert excel_bytes is not None

    # Load and verify workbook
    wb = openpyxl.load_workbook(excel_bytes)
    ws = wb.active

    # Check headers (row 1)
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "N°"
    assert headers[1] == "COLABORADOR"
    assert headers[-1] == "STATUS"

    # Check data row (row 2)
    row_data = [cell.value for cell in ws[2]]
    assert row_data[0] == 1  # N° (row number)
    assert row_data[1] == "João Silva"
    assert row_data[2] == "12345678901"
    assert row_data[-1] == "Ativo"


def test_generate_excel_empty_data():
    """Test Excel generation with empty data."""
    generator = EmployeeExcelGenerator()
    excel_bytes = generator.generate([])
    assert excel_bytes is not None

    wb = openpyxl.load_workbook(excel_bytes)
    ws = wb.active

    # Should have headers only
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "N°"

    # No data rows
    row2 = [cell.value for cell in ws[2]]
    assert all(v is None for v in row2)


def test_generate_excel_multiple_rows():
    """Test Excel generation with multiple data rows."""
    generator = EmployeeExcelGenerator()
    data = [
        {
            "employee": f"Employee {i}",
            "code": f"CPF{i}",
            "role": "Dev",
            "project": "P1",
            "bu": "TI",
            "cost_center": "CC",
            "cost_center_code": "CC01",
            "manager": "M",
            "executive": "E",
            "workload": "HO",
            "equipment_description": "Note",
            "patrimony": f"PAT-{i:03d}",
            "equipment_standard": "Std",
            "status": "Ativo",
        }
        for i in range(1, 4)
    ]

    excel_bytes = generator.generate(data)
    wb = openpyxl.load_workbook(excel_bytes)
    ws = wb.active

    # Verify row numbers
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=3, column=1).value == 2
    assert ws.cell(row=4, column=1).value == 3

    # Verify names
    assert ws.cell(row=2, column=2).value == "Employee 1"
    assert ws.cell(row=4, column=2).value == "Employee 3"


def test_generate_excel_header_formatting():
    """Test that headers are bold and centered."""
    generator = EmployeeExcelGenerator()
    excel_bytes = generator.generate([])

    wb = openpyxl.load_workbook(excel_bytes)
    ws = wb.active

    header_cell = ws.cell(row=1, column=1)
    assert header_cell.font.bold is True
    assert header_cell.alignment.horizontal == "center"
